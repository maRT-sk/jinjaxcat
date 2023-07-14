# Standard library imports
import csv
import io
import json
import os
import xml.dom.minidom
from collections import namedtuple
from io import BytesIO

# Third party imports
import chardet
import openpyxl
import pandas as pd
import requests
import streamlit as st
from jinja2 import FileSystemLoader, select_autoescape
from lxml import etree

# Local application/library specific imports
from utils.environment import CustomEnvironment

# Defining a named tuple to hold the result data
Result = namedtuple('Result', ['type', 'msg', 'log'])


def change_dict_keys(original_dict: dict, key_mapping: dict) -> dict:
    """
    Replaces the keys in a dictionary according to the provided mapping.
    Unmapped keys remain the same.

    :param original_dict: Original dictionary whose keys are to be replaced
    :param key_mapping: Dictionary having mapping {old_key : new_key}
    :return: A new dictionary with keys replaced.
    """
    new_dict = {}
    for key in original_dict:
        if key in key_mapping:
            new_key = key_mapping[key]
            new_dict[new_key] = original_dict[key]
        else:
            new_dict[key] = original_dict[key]
    return new_dict


@st.cache_data(show_spinner="Generating output...")
def generate_output(input_files: list, template_file: io.BytesIO, key_mapping: dict) -> bytes | str:
    """
    Function that generates a file from given input_files and a template_file.

    :param input_files: A list of input files containing data for the template.
    :param template_file: The template file.
    :param key_mapping: A dictionary that maps old keys to new keys.
    :return: A bytes object representing the rendered file.
    """
    template_bytes = template_file.getvalue()  # Get the bytes of the template file
    data_dict = load_data(input_files)  # Load the data from the input files into a dictionary
    data_dict = change_dict_keys(data_dict, key_mapping)  # Change the keys in the loaded data
    environment = create_environment()  # Create a custom Jinja2 environment

    # Check the file extension to decide how to render the template
    if template_file.name.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(filename=io.BytesIO(template_bytes))  # Load the workbook from the byte data
        for sheet_name in workbook.sheetnames:  # Iterate through each sheet in the workbook
            sheet = workbook[sheet_name]  # Access the sheet by its name

            # Iterate over each cell in each row of the sheet (limited to a maximum of 50 rows and 100 columns)
            for row in sheet.iter_rows(min_row=1, max_row=50, min_col=1, max_col=100):
                for cell in row:

                    # Check if the cell value starts with '{', indicating a Jinja2 template
                    if isinstance(cell.value, str) and cell.value.startswith('{'):

                        template = environment.from_string(cell.value)  # Create a Jinja2 template from the cell value
                        rendered_output = template.render(**data_dict)  # Render the template using the data dictionary
                        rendered_output = rendered_output.split('##')[0:-1]  # Split the rendered output by '##'

                        data = pd.Series(rendered_output)  # Convert the rendered output into a Pandas Series
                        series = data.apply(pd.to_numeric, errors='ignore')  # function that converts str to numeric

                        # Write the output values to the subsequent cells in the same column
                        for i, value in enumerate(series):
                            sheet.cell(row=cell.row + i, column=cell.column).value = value

        output_bytes = BytesIO()  # Create a BytesIO object to store the output data
        workbook.save(output_bytes)  # Save the workbook to the BytesIO object
        output_bytes.seek(0)  # Set the position of the BytesIO object to the beginning
        return output_bytes.getvalue()  # Return the byte representation of the Excel file
    # If the template file is not an Excel file, process it as a text-based file
    else:
        string_object = template_bytes.decode(chardet.detect(template_bytes)['encoding'])  # Decode bytes into a string
        template = environment.from_string(string_object)  # Create a Jinja2 template from the decoded string
        return template.render(**data_dict)  # Render the template with the data dictionary and return the result


@st.cache_data(show_spinner="Preparing data...")
def load_data(input_files: list) -> dict:
    """
    Function that loads data from various file types (CSV, Excel, JSON, and REST).
    The function returns a dictionary where each key-value pair corresponds to an input file and its contents.

    :param input_files: List of strings representing file paths of the input files.
    :return: Dictionary where each key-value pair corresponds to an input file and its contents.
    """
    data_dict = {}  # Initialize a dictionary to store the data

    # Loop over each input file
    for file in input_files:

        extension = os.path.splitext(file.name)[-1]  # Get the extension name
        name = file.name.replace('.','_')  # Get the filename

        if name in data_dict:
            raise Exception(
                f"Duplicate Detected: The file '{name}' already exists. Please rename your input files.")

        if extension == '.csv':
            # Get the bytes object of the file and decode it using the detected encoding
            bytes_object = file.getvalue()
            string_object = bytes_object.decode(chardet.detect(bytes_object)['encoding'])
            # dialect = csv.Sniffer().sniff(string_object)

            # Load the CSV file into a pandas DataFrame and convert it to a dictionary
            df = pd.read_csv(io.StringIO(string_object), dtype=str, quoting=3, sep=None, engine="python").fillna('')
            data_dict[name] = df.to_dict('records')

        elif extension == '.xlsx':
            # Loop over each sheet in the Excel file
            for sheet in pd.ExcelFile(file).sheet_names:
                df = pd.read_excel(file, sheet, engine='openpyxl', dtype=str).fillna('')
                # Load the sheet into a pandas DataFrame and convert it to a dictionary
                data_dict[f"{sheet}_{name}"] = df.to_dict('records')

        elif extension == '.rest':
            # Get the bytes object of the file and decode and extract the HTTP method and headers
            bytes_object = file.getvalue()
            string_object = bytes_object.decode(chardet.detect(bytes_object)['encoding'])
            lines = string_object.split('\n')  # Split the decoded string into lines
            method, url = lines[0].split()  # Extract the HTTP method and URL from the first line
            headers = {line.split(": ")[0]: line.split(": ")[1].strip() for line in lines[1:] if ": " in line}
            if method.lower() == 'get':  # If the HTTP method is GET, send a GET request to the URL with the headers
                response = requests.get(url, headers=headers)
            else:
                raise ValueError(f"Unsupported HTTP method: {method}")
            # If the response status code is 2xx, convert the JSON response to a  DataFrame and then to a dictionary
            if 200 <= response.status_code < 300:  # Status code is 2xx
                data_dict[name] = response.json()
            else:
                response.raise_for_status()  # If the response status code is not 2xx, raise an exception

        elif extension == '.json':
            # Load the JSON file into a pandas DataFrame and convert it to a dictionary
            data_dict[name] = json.loads(file.getvalue())



    return data_dict  # Return the dictionary containing all the data


def create_environment() -> CustomEnvironment:
    """
    Create a custom Jinja2 environment.
    :return: Custom Jinja2 environment object.
    """
    # Initialize a custom Jinja2 environment
    environment = CustomEnvironment(loader=FileSystemLoader(''), autoescape=select_autoescape(["html", "htm", "xml"]))
    return environment


def prettify_output(content: str, extension) -> str | None:
    """
    Format the content based on the file type (.xml or .csv).
    If the file is neither .xml nor .csv, the content is returned as is.

    :param content: String containing the content to be formatted.
    :param output_file_name: String containing the name of the output file.
    :return: str: The formatted content.
    """

    # If the file is an XML file, create a prettified version of the XML with indentations and line breaks
    if extension == '.xml':
        try:
            dom = xml.dom.minidom.parseString(content)  # Parse the XML content
            prettified_xml = dom.toprettyxml(indent="  ", newl="\n")  # Create a prettified XML with indentation, breaks
            lines = [line for line in prettified_xml.split("\n") if line.strip()]  # Erase lines containing white spaces
            prettified_content = "\n".join(lines)  # Join the lines into a single string
            return prettified_content
        except:  # noqa: E722
            # If the XML content is not valid, return None
            # Note: the exception handling here is not specific and should be improved in future revisions
            return None

    # If the file is a CSV file, format the CSV content with aligned columns
    elif extension == '.csv':
        try:
            rows = content.strip().split("\n")  # Split the CSV content into rows
            dialect = csv.Sniffer().sniff(rows[0])  # Detect the delimiter from the header row
            delimiter = dialect.delimiter
            max_columns = max(len(row.split(delimiter)) for row in rows)  # Determine the maximum number of columns

            formatted_rows = []
            # Split each row into columns, trim leading and trailing whitespace, and pad with empty values if necessary
            for row in rows:
                columns = [column.strip() for column in row.split(delimiter)]
                columns += [""] * (max_columns - len(columns))
                formatted_rows.append(columns)

            # Join the formatted rows into a single string
            formatted_csv = "\n".join(
                delimiter.join(row) for row in formatted_rows if any(column.strip() for column in row)
            )
            return formatted_csv
        except:  # noqa: E722
            # If the CSV content is not valid or cannot be parsed, return None
            # Note: the exception handling here is not specific and should be improved in future revisions
            return None
    return None  # If the file is neither .xml nor .csv, return None


def validate_xml(xml_file, schema_path) -> Result:
    """
    Validate an XML file against a DTD or XSD schema.
    :param xml_file: String containing the XML file content.
    :param schema_path: String containing the path to the schema file.
    :return: Instance of Result with status, title and message.
    """

    schema_type = schema_path[schema_path.rfind("."):]  # Extract schema type by retrieving the file extension
    try:
        xml_doc = etree.fromstring(xml_file)  # Parse the XML file
    except etree.XMLSyntaxError as e:
        # If XML parsing fails, return Result with failure status and error details
        return Result("KO", e.__class__.__name__, e)

    if schema_type == '.dtd':  # Validate the XML against the appropriate schema based on the schema type
        try:
            schema = etree.DTD(file=schema_path)  # Load the DTD schema
            is_valid = schema.validate(xml_doc)  # Validate the XML document against the schema
            errors = schema.error_log.filter_from_errors()  # Get the list of validation errors
        except (etree.DTDParseError, etree.XMLSyntaxError) as e:
            # If DTD parsing or validation fails, return Result with failure status and error details
            return Result("KO", e.__class__.__name__, e)
    elif schema_type == '.xsd':
        try:
            schema_doc = etree.parse(schema_path)  # Load the XSD schema
            schema = etree.XMLSchema(schema_doc)
            is_valid = schema.validate(xml_doc)  # Validate the XML document against the schema
            errors = schema.error_log  # Get the list of validation errors
        except etree.XMLSyntaxError as e:
            # If XSD parsing or validation fails, return Result with failure status and error details
            return Result("KO", e.__class__.__name__, e)
    else:
        return Result("KO", "Error", "The schema type is not recognize")  # If the schema type is not recognized

    # Log whether the XML is valid or not, and any errors
    if is_valid:
        return Result("OK", "Validation Passed",
                      f"The XML output successfully matches the specified ({schema_type.upper()}) schema.")
    else:
        return Result("KO", "Validation Failed",
                      f"The XML output does not conform to the specified ({schema_type.upper()}) schema: \n\n{errors}")
