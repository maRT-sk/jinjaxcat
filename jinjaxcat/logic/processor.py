# Standard library imports
import csv
import json
import os
import traceback
import xml.dom.minidom

# Third party imports
import chardet
import jinja2
from jinja2 import FileSystemLoader, select_autoescape
from lxml import etree
from openpyxl import load_workbook

# Local application/library specific imports
from jinjaxcat.logic.enviroment import CustomEnvironment, log


def load_data(input_files: list) -> dict:
    """
    Load data from CSV, Excel and JSON files located in a specified input folder.
    :param input_files: List of strings representing file paths of the input files.
    :return: Dictionary where each key-value pair corresponds to an input file and its contents.
    """
    data_dict = {}  # Initialize a dictionary to store the data

    # Loop over each input file
    for file_path in input_files:
        file_name = os.path.basename(file_path)  # Extract file name from file path
        encoding = detect_encoding(file_path)  # Detect file encoding
        # Check if the file is a CSV file and write its contents into a dictionary
        if file_name.endswith('.csv'):
            with open(file_path, newline='', encoding=encoding) as csvfile:
                sample = csvfile.read(4096)  # Read a larger sample of the file
                dialect = csv.Sniffer().sniff(sample)
                csvfile.seek(0)  # Reset the file pointer to the beginning
                reader = csv.DictReader(csvfile, delimiter=str(dialect.delimiter))
                data_dict[file_name.split('.')[0]] = list(reader)
        # Check if the file is a JSON file and write its contents into a dictionary
        elif file_name.endswith('.json'):
            with open(file_path, encoding=encoding) as file:
                json_data = file.read()
                dictionary = json.loads(json_data)
                data_dict[f"{file_name.split('.')[0]}"] = dictionary
        # Check if the file is an Excel file and write its contents into a dictionary
        elif file_name.endswith('.xlsx'):
            # Load workbook in read_only mode for efficient reading of large files
            wb = load_workbook(filename=file_path, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                data = []
                keys = None
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:  # The first row is assumed to be the header
                        keys = row
                    else:
                        # For each row, create a dictionary mapping column name to cell value
                        data.append(dict(zip(keys, [str(cell) if cell is not None else '' for cell in row])))
                # Assign each sheet's data to a separate key in the data dictionary
                # Key: Filename_Sheetname, Value: List of rows as dictionaries
                data_dict[f"{file_name.split('.')[0]}_{sheet}"] = data

    return data_dict  # Return the data dictionary


def create_environment() -> CustomEnvironment:
    """
    Create a custom Jinja2 environment.
    :return: Custom Jinja2 environment object.
    """
    # Initialize a custom Jinja2 environment
    environment = CustomEnvironment(
        loader=FileSystemLoader(''),
        autoescape=select_autoescape(["html", "htm", "xml"])
    )
    return environment


def prettify_output(content: str, output_file_name: str) -> str | None:
    """
    Format the content based on the file type (.xml or .csv).
    If the file is neither .xml nor .csv, the content is returned as is.

    :param content: String containing the content to be formatted.
    :param output_file_name: String containing the name of the output file.
    :return: str: The formatted content.
    """
    # If the file is an XML file, create a prettified version of the XML with indentations and line breaks
    if output_file_name.endswith(".xml"):
        # Parse the XML content
        try:
            dom = xml.dom.minidom.parseString(content)
            # Create a prettified version of the XML with indentations and line breaks
            prettified_xml = dom.toprettyxml(indent="  ", newl="\n")
            # Remove lines that only contain white spaces
            lines = [line for line in prettified_xml.split("\n") if line.strip()]
            # Join the lines into a single string
            prettified_content = "\n".join(lines)
            return prettified_content
        except:  # needs to be handled better some day  # noqa: E722
            # If the XML content is not valid, return None
            return None

    # If the file is a CSV file, format the CSV content with aligned columns
    elif output_file_name.endswith(".csv"):
        try:
            rows = content.strip().split("\n")  # Split the CSV content into rows
            dialect = csv.Sniffer().sniff(rows[0])  # Detect the delimiter from the header row
            delimiter = dialect.delimiter
            max_columns = max(len(row.split(delimiter)) for row in rows)  # Determine the maximum number of columns

            formatted_rows = []
            # Split each row into columns, trim leading and trailing whitespace, and pad with empty values if necessary
            for row in rows:
                columns = [column.strip() for column in row.split(delimiter)]
                columns += [""] * (max_columns - len(columns))
                formatted_rows.append(columns)

            # Join the formatted rows into a single string
            formatted_csv = "\n".join(
                delimiter.join(row) for row in formatted_rows if any(column.strip() for column in row)
            )
            return formatted_csv
        except:  # needs to be handled better some day  # noqa: E722
            return None
    return None  # If the file is neither .xml nor .csv, return None


def render_template(data_dict: dict, environment: CustomEnvironment, template_name: str, output_file_name: str,
                    is_prettify_output: bool) -> str | None:
    """
    Render a Jinja2 template using the loaded data and save the rendered output into a file.
    :param data_dict: Dictionary containing loaded data.
    :param environment: Custom Jinja2 environment object.
    :param template_name: String representing the name of the template file.
    :param output_file_name: String representing the name of the output file.
    :param is_prettify_output: Boolean indicating whether to prettify the output file.
    :return: String containing the name of the output file.
    """

    # Get the Jinja2 template from the environment
    try:
        template = environment.get_template(template_name)
    except (jinja2.exceptions.TemplateSyntaxError, jinja2.exceptions.TemplateNotFound) as e:
        log(f"Traceback:\n {''.join(traceback.format_tb(e.__traceback__)[-1:])}", True)  # NOQA
        log(f"{e.__class__.__name__}: {e} -> see traceback for more details", True)  # NOQA
        return None

    # Render the template using the data dictionary
    try:
        output = template.render(data_dict)
    except (jinja2.exceptions.UndefinedError, TypeError) as e:
        tracebacks = traceback.format_tb(e.__traceback__)[-2:]  # Retrieve the last two tracebacks
        log(f"Traceback:\n{''.join(tracebacks)}\n{e.__class__.__name__}: {e}")  # NOQA
        log(f"{e.__class__.__name__}: {e} -> see traceback for more details", True)  # NOQA
        return None

    # If the output should be prettified, prettify it
    if is_prettify_output:
        # Prettify the output and check if it was successful
        output = prettify_output(output, output_file_name)
        if output is None:
            # If an unexpected error occurred during prettifying, log a message
            log("An unexpected error occurred. Please try running it without prettifying.", True)  # NOQA
            return None

    # Write the output to a file
    with open(output_file_name, 'w', encoding='utf-8-sig') as f:
        f.write(output)
    return output_file_name


def validate_xml(xml_path: str, schema_path: str, schema_type: str) -> bool:
    """
     Validate an XML file against a DTD or XSD schema.
     :param xml_path: String containing the path to the XML file.
     :param schema_path: String containing the path to the schema file.
     :param schema_type: String indicating the type of the schema (DTD or XSD).
     :return: Boolean indicating whether the XML is valid or not.
     """
    try:
        xml_doc = etree.parse(xml_path)  # Parse the XML file
    except etree.XMLSyntaxError as e:
        log(f"{e.__class__.__name__}: {e}", True)  # NOQA
        return False

    # Validate the XML against the appropriate schema based on the schema type
    if schema_type == 'dtd':
        try:
            schema = etree.DTD(file=schema_path)
            is_valid = schema.validate(xml_doc)
            errors = schema.error_log.filter_from_errors()
        except (etree.DTDParseError, etree.XMLSyntaxError) as e:
            log(f"{e.__class__.__name__}: {e}", True)  # NOQA
            return False
    elif schema_type == 'xsd':
        try:
            schema_doc = etree.parse(schema_path)
            schema = etree.XMLSchema(schema_doc)
            is_valid = schema.validate(xml_doc)
            errors = schema.error_log
        except etree.XMLSyntaxError as e:
            log(f"{e.__class__.__name__}: {e}", True)  # NOQA
            return False
    else:
        return False

    # Log whether the XML is valid or not, and any errors
    if is_valid:
        msg = f"Valid XML: Output matches ({schema_type.upper()}) schema."
        log(msg)  # noqa
        return True
    else:
        msg = f"Invalid XML: Does not conform specified ({schema_type.upper()}) schema."
        log(msg)  # noqa
        for e in errors:
            log(str(e))  # noqa
        return False


def validate_json(file_path: str) -> bool:
    """
    Validate a JSON file. If the file is invalid, log the details of the error.
    :param file_path: String containing the path to the JSON file.
    :return: Boolean indicating whether the JSON file is valid or not.
    """
    # Open the JSON file and try to parse it
    try:
        with open(file_path, encoding='utf-8-sig') as f:
            json.load(f)
    except ValueError as e:
        # If a ValueError occurs, it means the JSON is invalid
        log(f'Invalid JSON file: {file_path}')  # noqa
        log(f'Details: {e}')  # noqa
        return False  # Return False since the JSON file is invalid
    log(f'Valid JSON file: {file_path}')  # noqa
    return True  # Return True since the JSON file is valid


def validate_output(output_file: str, validation_type: str, validate_against: str) -> bool:
    """
    Validate the output file if a validation schema is provided.
    :param output_file: String containing the name of the output file.
    :param validation_type: String indicating the type of validation (XML or JSON).
    :param validate_against: String containing the path to the validation schema file.
    :return: Boolean indicating whether the output file is valid or not.
    """

    # If it's XML, we need to determine the schema type and then validate the XML file against the schema
    if validation_type == "xml":
        try:
            schema_type = validate_against.split(".")[-1]  # Attempt to extract the schema type
        except AttributeError:
            log(f'Incorrect schema type', True)  # noqa
            return False
        if schema_type in ['dtd', 'xsd']:
            if validate_xml(output_file, validate_against, schema_type):
                return True
            else:
                return False
        else:
            log(f'Invalid file extension for validation: {schema_type}')  # noqa
            raise ValueError(f"Invalid file extension for validation: {schema_type}")
    # If the validation type is JSON, validate the JSON file
    elif validation_type == "json":
        if validate_json(output_file):
            return True
        else:
            return False


def detect_encoding(file_path: str) -> str:
    """
    Detect the encoding of a file.
    :param file_path: String containing the path to the file.
    :return: String representing the detected encoding.
    """
    with open(file_path, 'rb') as f:  # Open the file in binary mode
        result = chardet.detect(f.read())  # Use chardet to detect the file's encoding
    return result['encoding']  # Return the detected encoding
